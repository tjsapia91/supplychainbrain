"""
Document parsers for shipping documents.
Extracts container-level and line-item data from:
- Excel files (packing lists, commercial invoices)
- PDFs (packing lists, commercial invoices, BOLs)
- Images (photos of documents via OCR)
"""
import re
import io
import json
from decimal import Decimal, InvalidOperation
from datetime import datetime, date


# ─── Column header mapping ───────────────────────────────────────
# Maps common header variations to our field names
HEADER_MAP = {
    # Item identification
    'item': 'item_no', 'item_no': 'item_no', 'item #': 'item_no', 'item#': 'item_no',
    'item number': 'item_no', 'upc': 'item_no', 'sku': 'item_no', 'sku#': 'item_no',
    'product code': 'item_no', 'part number': 'item_no', 'part no': 'item_no',
    'product no': 'item_no', 'barcode': 'item_no',

    # Description
    'description': 'description', 'desc': 'description', 'product': 'description',
    'item description': 'description', 'product name': 'description',
    'product description': 'description', 'name': 'description', 'goods': 'description',

    # Quantity
    'qty': 'quantity', 'quantity': 'quantity', 'units': 'quantity', 'pcs': 'quantity',
    'pieces': 'quantity', 'total qty': 'quantity', 'total quantity': 'quantity',
    'qty shipped': 'quantity', 'ship qty': 'quantity', 'order qty': 'quantity',

    # Cartons
    'cartons': 'cartons', 'cases': 'cartons', 'ctns': 'cartons', 'ctn': 'cartons',
    'no of cartons': 'cartons', 'no. of cartons': 'cartons', 'boxes': 'cartons',
    'total cartons': 'cartons', 'ttl ctns': 'cartons',

    # CBM
    'cbm': 'cbm', 'cubic meter': 'cbm', 'cubic meters': 'cbm', 'volume': 'cbm',
    'total cbm': 'cbm', 'vol': 'cbm', 'm3': 'cbm', 'cu.m': 'cbm',

    # Weight
    'weight': 'total_weight', 'gross weight': 'total_weight', 'gw': 'total_weight',
    'total weight': 'total_weight', 'weight (kg)': 'total_weight', 'g.w.': 'total_weight',
    'gross wt': 'total_weight', 'net weight': 'total_weight', 'nw': 'total_weight',
    'wt': 'total_weight', 'weight kg': 'total_weight',

    # Value
    'value': 'line_value', 'amount': 'line_value', 'total': 'line_value',
    'total value': 'line_value', 'unit price': 'unit_price', 'price': 'unit_price',
    'fob value': 'line_value', 'invoice value': 'line_value',

    # HTS
    'hts': 'hts_code', 'hts code': 'hts_code', 'hs code': 'hts_code',
    'tariff': 'hts_code', 'harmonized code': 'hts_code', 'hts#': 'hts_code',
    'hs#': 'hts_code', 'customs code': 'hts_code',

    # Destination
    'destination': 'destination', 'dest': 'destination', 'ship to': 'destination',
    'channel': 'destination', 'warehouse': 'destination', 'delivery to': 'destination',
}

# Key-value patterns for container-level fields
CONTAINER_PATTERNS = {
    'container_number': [
        r'container\s*(?:#|no\.?|number)?\s*[:\-]?\s*([A-Z]{4}\d{7})',
        r'([A-Z]{4}\d{7})',  # standard container number format
    ],
    'hbl_number': [
        r'(?:hbl|h\.?b\.?l\.?|house\s*b/?l)\s*(?:#|no\.?|number)?\s*[:\-]?\s*(\S+)',
        r'(?:bill\s*of\s*lading)\s*(?:#|no\.?)?\s*[:\-]?\s*(\S+)',
    ],
    'booking_reference': [
        r'(?:booking|bkg)\s*(?:#|no\.?|ref\.?)?\s*[:\-]?\s*(\S+)',
    ],
    'commercial_invoice': [
        r'(?:commercial\s*)?invoice\s*(?:#|no\.?|number)?\s*[:\-]?\s*(\S+)',
        r'(?:ci|inv)\s*(?:#|no\.?)?\s*[:\-]?\s*(\S+)',
    ],
    'forwarder': [
        r'(?:forwarder|freight\s*forwarder|shipper)\s*[:\-]?\s*(.+?)(?:\n|$)',
    ],
    'vessel': [
        r'(?:vessel|ship|v/v)\s*(?:name)?\s*[:\-]?\s*(.+?)(?:\n|$)',
    ],
    'port_of_loading': [
        r'(?:port\s*of\s*(?:loading|origin)|pol|from)\s*[:\-]?\s*(.+?)(?:\n|$)',
    ],
    'port_of_discharge': [
        r'(?:port\s*of\s*(?:discharge|destination)|pod|to)\s*[:\-]?\s*(.+?)(?:\n|$)',
    ],
}

DATE_PATTERNS = {
    'date_sailed': [
        r'(?:sail(?:ed|ing)?\s*date|departure|etd|date\s*sailed)\s*[:\-]?\s*(\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4})',
    ],
    'eta_port': [
        r'(?:eta|arrival|estimated\s*arrival)\s*[:\-]?\s*(\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4})',
    ],
}


def normalize_header(header):
    """Normalize a column header to a known field name."""
    if not header:
        return None
    h = str(header).strip().lower().replace('_', ' ').replace('.', ' ').strip()
    # Direct match
    if h in HEADER_MAP:
        return HEADER_MAP[h]
    # Partial match
    for key, field in HEADER_MAP.items():
        if key in h or h in key:
            return field
    return None


def safe_int(val):
    """Safely convert a value to int."""
    if val is None:
        return 0
    try:
        # Handle strings with commas
        if isinstance(val, str):
            val = val.replace(',', '').strip()
        return int(float(val))
    except (ValueError, TypeError):
        return 0


def safe_decimal(val):
    """Safely convert a value to Decimal."""
    if val is None:
        return Decimal('0')
    try:
        if isinstance(val, str):
            val = val.replace(',', '').replace('$', '').strip()
        return Decimal(str(val)).quantize(Decimal('0.0001'))
    except (InvalidOperation, ValueError, TypeError):
        return Decimal('0')


def parse_date(text):
    """Try to parse a date from various formats."""
    if not text:
        return None
    if isinstance(text, (date, datetime)):
        return text if isinstance(text, date) else text.date()
    text = str(text).strip()
    formats = [
        '%m/%d/%Y', '%m-%d-%Y', '%m.%d.%Y',
        '%d/%m/%Y', '%d-%m-%Y', '%d.%m.%Y',
        '%Y-%m-%d', '%Y/%m/%d',
        '%m/%d/%y', '%m-%d-%y',
        '%b %d, %Y', '%B %d, %Y',
        '%d %b %Y', '%d %B %Y',
    ]
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def extract_container_fields(text):
    """Extract container-level fields from free text using regex patterns."""
    fields = {}
    text_upper = text.upper()

    # Container-level string fields
    for field_name, patterns in CONTAINER_PATTERNS.items():
        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                fields[field_name] = match.group(1).strip()
                break

    # Date fields
    for field_name, patterns in DATE_PATTERNS.items():
        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                parsed = parse_date(match.group(1))
                if parsed:
                    fields[field_name] = parsed.isoformat()
                break

    # Incoterms detection
    for term in ['DDP', 'FOB', 'CIF', 'EXW', 'FCA', 'CFR', 'CIP', 'DAP', 'DPU']:
        if term in text_upper:
            fields['incoterms'] = term
            break

    # Container type detection
    if '40HQ' in text_upper or '40HC' in text_upper or "40' HIGH CUBE" in text_upper:
        fields['container_type'] = '40hq'
    elif '40FT' in text_upper or "40'" in text_upper or '40 FT' in text_upper:
        fields['container_type'] = '40ft'
    elif '20FT' in text_upper or "20'" in text_upper or '20 FT' in text_upper:
        fields['container_type'] = '20ft'

    return fields


# ─── Excel Parser ─────────────────────────────────────────────────

def parse_excel(file_obj):
    """
    Parse an Excel file (packing list, commercial invoice, etc.)
    Returns: {'container': {field: value}, 'items': [{field: value}, ...], 'raw_text': str}
    """
    import openpyxl

    wb = openpyxl.load_workbook(file_obj, data_only=True)
    result = {'container': {}, 'items': [], 'raw_text': '', 'doc_type': 'excel'}

    # Try each sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Collect all text for container-level extraction
        all_text = []
        for row in rows:
            for cell in row:
                if cell is not None:
                    all_text.append(str(cell))
        sheet_text = '\n'.join(all_text)
        result['raw_text'] += sheet_text + '\n'

        # Find the header row (row with the most recognized column names)
        best_header_row = -1
        best_header_count = 0
        best_mapping = {}

        for i, row in enumerate(rows):
            if not row or all(c is None for c in row):
                continue
            mapping = {}
            count = 0
            for j, cell in enumerate(row):
                field = normalize_header(cell)
                if field:
                    mapping[j] = field
                    count += 1
            if count > best_header_count and count >= 2:
                best_header_count = count
                best_header_row = i
                best_mapping = mapping

        if best_header_row >= 0:
            # Extract data rows after header
            for row in rows[best_header_row + 1:]:
                if not row or all(c is None for c in row):
                    continue
                item = {}
                has_data = False
                for col_idx, field_name in best_mapping.items():
                    if col_idx < len(row):
                        val = row[col_idx]
                        if val is not None and str(val).strip():
                            has_data = True
                            if field_name in ('quantity', 'cartons'):
                                item[field_name] = safe_int(val)
                            elif field_name in ('cbm', 'total_weight', 'line_value', 'unit_price'):
                                item[field_name] = str(safe_decimal(val))
                            else:
                                item[field_name] = str(val).strip()
                if has_data and (item.get('item_no') or item.get('description')):
                    # Calculate line_value from unit_price if needed
                    if 'unit_price' in item and 'line_value' not in item and item.get('quantity'):
                        price = safe_decimal(item.pop('unit_price'))
                        qty = safe_int(item.get('quantity', 0))
                        item['line_value'] = str(price * qty)
                    elif 'unit_price' in item:
                        item.pop('unit_price', None)
                    result['items'].append(item)

    # Extract container-level fields from all text
    result['container'] = extract_container_fields(result['raw_text'])

    return result


# ─── PDF Parser ───────────────────────────────────────────────────

def parse_pdf(file_obj):
    """
    Parse a PDF file using pdfplumber for text and table extraction.
    Returns: {'container': {field: value}, 'items': [{field: value}, ...], 'raw_text': str}
    """
    import pdfplumber

    result = {'container': {}, 'items': [], 'raw_text': '', 'doc_type': 'pdf'}

    with pdfplumber.open(file_obj) as pdf:
        all_text = []
        all_tables = []

        for page in pdf.pages:
            # Extract text
            text = page.extract_text()
            if text:
                all_text.append(text)

            # Extract tables
            tables = page.extract_tables()
            for table in tables:
                if table:
                    all_tables.append(table)

        result['raw_text'] = '\n'.join(all_text)

        # Extract container-level fields from text
        result['container'] = extract_container_fields(result['raw_text'])

        # Parse tables for line items
        for table in all_tables:
            if len(table) < 2:
                continue

            # First row is likely headers
            header_row = table[0]
            mapping = {}
            for j, cell in enumerate(header_row):
                field = normalize_header(cell)
                if field:
                    mapping[j] = field

            if len(mapping) >= 2:
                for row in table[1:]:
                    if not row or all(c is None or str(c).strip() == '' for c in row):
                        continue
                    item = {}
                    has_data = False
                    for col_idx, field_name in mapping.items():
                        if col_idx < len(row):
                            val = row[col_idx]
                            if val is not None and str(val).strip():
                                has_data = True
                                if field_name in ('quantity', 'cartons'):
                                    item[field_name] = safe_int(val)
                                elif field_name in ('cbm', 'total_weight', 'line_value', 'unit_price'):
                                    item[field_name] = str(safe_decimal(val))
                                else:
                                    item[field_name] = str(val).strip()
                    if has_data and (item.get('item_no') or item.get('description')):
                        if 'unit_price' in item and 'line_value' not in item and item.get('quantity'):
                            price = safe_decimal(item.pop('unit_price'))
                            qty = safe_int(item.get('quantity', 0))
                            item['line_value'] = str(price * qty)
                        elif 'unit_price' in item:
                            item.pop('unit_price', None)
                        result['items'].append(item)

    return result


# ─── Image Parser (OCR) ──────────────────────────────────────────

def parse_image(file_obj):
    """
    Parse an image file using Tesseract OCR.
    Returns: {'container': {field: value}, 'items': [{field: value}, ...], 'raw_text': str}
    """
    from PIL import Image
    result = {'container': {}, 'items': [], 'raw_text': '', 'doc_type': 'image'}

    try:
        import pytesseract
        img = Image.open(file_obj)
        # Improve OCR quality
        img = img.convert('L')  # Grayscale
        text = pytesseract.image_to_string(img)
        result['raw_text'] = text
        result['container'] = extract_container_fields(text)
    except ImportError:
        result['raw_text'] = '[OCR not available - pytesseract not installed]'
    except Exception as e:
        result['raw_text'] = f'[OCR error: {str(e)}]'

    return result


# ─── Main entry point ─────────────────────────────────────────────

def parse_document(file_obj, filename):
    """
    Auto-detect file type and parse accordingly.
    Returns: {'container': {}, 'items': [], 'raw_text': '', 'doc_type': ''}
    """
    ext = filename.lower().rsplit('.', 1)[-1] if '.' in filename else ''

    if ext in ('xlsx', 'xls', 'xlsm'):
        return parse_excel(file_obj)
    elif ext == 'pdf':
        return parse_pdf(file_obj)
    elif ext in ('jpg', 'jpeg', 'png', 'gif', 'bmp', 'tiff', 'tif', 'webp'):
        return parse_image(file_obj)
    elif ext == 'csv':
        return parse_csv(file_obj)
    else:
        # Try as Excel first, then PDF, then image
        try:
            return parse_excel(file_obj)
        except Exception:
            file_obj.seek(0)
            try:
                return parse_pdf(file_obj)
            except Exception:
                file_obj.seek(0)
                return parse_image(file_obj)


def parse_csv(file_obj):
    """Parse a CSV file for packing list data."""
    import csv

    result = {'container': {}, 'items': [], 'raw_text': '', 'doc_type': 'csv'}

    content = file_obj.read()
    if isinstance(content, bytes):
        content = content.decode('utf-8-sig')

    lines = content.splitlines()
    result['raw_text'] = content

    reader = csv.reader(lines)
    rows = list(reader)
    if len(rows) < 2:
        return result

    # Map headers
    header_row = rows[0]
    mapping = {}
    for j, cell in enumerate(header_row):
        field = normalize_header(cell)
        if field:
            mapping[j] = field

    if mapping:
        for row in rows[1:]:
            if not row or all(not c.strip() for c in row):
                continue
            item = {}
            has_data = False
            for col_idx, field_name in mapping.items():
                if col_idx < len(row):
                    val = row[col_idx].strip()
                    if val:
                        has_data = True
                        if field_name in ('quantity', 'cartons'):
                            item[field_name] = safe_int(val)
                        elif field_name in ('cbm', 'total_weight', 'line_value'):
                            item[field_name] = str(safe_decimal(val))
                        else:
                            item[field_name] = val
            if has_data and (item.get('item_no') or item.get('description')):
                result['items'].append(item)

    result['container'] = extract_container_fields(result['raw_text'])
    return result
